from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.conf import settings
from django.db.models import Count, Sum, Q
from django.utils.timezone import now
from django.views.decorators.cache import never_cache
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,Paragraph, KeepInFrame)


from accounts.models import Result


from accounts.models import Result

from .forms import (
    LoginForm, RegisterForm, StudentForm,
    ItemForm, RegistrationForm
)
from .models import (
    AppealNotification, Student, Item, Registration,
    College, Team, Result,
    CustomUser, SiteSetting
)
from .utils import calculate_points

#homeeeeee
def home(request):
    return render(request, "accounts/home.html")

#👤 USER REGISTER
def user_register(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Account created successfully!")
            return redirect("login")
    else:
        form = RegisterForm()
    return render(request, "accounts/register.html", {"form": form})


#🔐 LOGIN
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect


def user_login(request):
    """
    Single source of truth for login.
    DO NOT create any other login() view.
    """
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()

        if not username or not password:
            messages.error(request, "Both username and password are required.")
            return render(request, "accounts/login.html")

        user = authenticate(
            request,
            username=username,
            password=password,
        )

        if user is not None:
            if not user.is_active:
                messages.error(request, "Your account is disabled.")
                return render(request, "accounts/login.html")

            login(request, user)

            # 🔀 Role-based redirect
            if user.role == "college":
                return redirect("college_dashboard")
            elif user.role == "organizer":
                return redirect("organizer_dashboard")
            elif user.role == "admin":
                return redirect("admin_dashboard")

            return redirect("home")

        messages.error(request, "Invalid username or password.")

    return render(request, "accounts/login.html")



#🚪 LOGOUT
def user_logout(request):
    logout(request)
    request.session.flush()
    return redirect("login")


#🎓 COLLEGE DASHBOARD
@login_required
def college_dashboard(request):
    if request.user.role != "college":
        return redirect("home")

    college = get_object_or_404(College, user=request.user)
    setting = SiteSetting.objects.first()

    # 🔐 If registration is CLOSED → redirect to summary
    if not setting or not setting.allow_student_registration:
        return redirect("student_summary")

    # ✅ Otherwise allow registration dashboard
    if request.method == "POST" and "add_student" in request.POST:
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save(commit=False)
            student.college = college
            student.save()
            messages.success(request, "Student added successfully.")
            return redirect("register_student")
    else:
        form = StudentForm()

    return render(request, "accounts/register.html", {
        "student_form": form,
        "students": Student.objects.filter(college=college),
        "items": Item.objects.all(),
        "college": college,
    })


from .decorators import registration_open_required


#🧑‍🎓 REGISTER STUDENT
@login_required
@registration_open_required
def register_student(request):
    if request.user.role != "college":
        return redirect("home")

    college = get_object_or_404(College, user=request.user)

    if request.method == "POST":
        Student.objects.create(
            college=college,
            name=request.POST["name"],
            id_card=request.POST["id_card"],
            date_of_birth=request.POST["date_of_birth"],
            department=request.POST["department"],
            year_of_joining=request.POST["year_of_joining"],
            current_year=request.POST["current_year"],
            id_card_file=request.FILES.get("id_card_file"),
        )
        messages.success(request, "Student registered")
        return redirect("register_student")

    return render(request, "accounts/register.html", {
        "students": Student.objects.filter(college=college),
        "college": college,
    })


#✏️ EDIT STUDENT
@login_required
@registration_open_required
def edit_student(request, student_id):
    student = get_object_or_404(Student, id=student_id)

    if request.method == "POST":
        for field in [
            "name", "id_card", "date_of_birth",
            "department", "current_year", "year_of_joining"
        ]:
            setattr(student, field, request.POST.get(field))

        if request.FILES.get("id_card_file"):
            student.id_card_file = request.FILES["id_card_file"]

        student.save()
        messages.success(request, "Student updated")
        return redirect("register_student")

    return render(request, "accounts/edit_student.html", {"student": student})





#👥 TEAM CREATION (FIXED)
@login_required
@registration_open_required
def team_creation(request):
    if request.user.role != "college":
        return redirect("home")

    college = get_object_or_404(College, user=request.user)

    students = Student.objects.filter(college=college)
    items = Item.objects.all()
    categories = Item.objects.values_list("category", flat=True).distinct()

    existing_item_ids = Team.objects.filter(
        college=college
    ).values_list("item_id", flat=True)

    if request.method == "POST":
        item = get_object_or_404(Item, id=request.POST.get("item"))
        selected_students = request.POST.getlist("students")

        if Team.objects.filter(college=college, item=item).exists():
            messages.error(request, "Team already exists")
            return redirect("team_creation")

        team = Team.objects.create(
            college=college,
            item=item,
            category=item.category,
        )
        team.students.set(selected_students)
        messages.success(request, "Team created")
        return redirect("team_creation")

    return render(request, "team/create_team.html", {
        "students": students,
        "items": items,
        "categories": categories,
        "teams": Team.objects.filter(college=college),
        "existing_item_ids": existing_item_ids,
        "college": college,
    })


#edit team
@login_required
@registration_open_required
def edit_team(request, team_id):
    if request.user.role != "college":
        messages.error(request, "Only college users can edit teams.")
        return redirect("home")

    college = get_object_or_404(College, user=request.user)
    team = get_object_or_404(Team, id=team_id, college=college)
    item = team.item

    if request.method == "POST":
        selected_students = request.POST.getlist("edit_students")

        if not selected_students:
            messages.error(request, "A team must have at least one participant.")
            return redirect("team_creation")

        if len(selected_students) > item.max_participants:
            messages.error(
                request,
                f"Maximum {item.max_participants} participants allowed for '{item.name}'."
            )
            return redirect("team_creation")

        single_limit = 4
        group_limit = 2

        for student_id in selected_students:
            student = Student.objects.get(id=student_id)

            single_count = (
                Team.objects.filter(
                    college=college,
                    students=student,
                    item__item_type="single",
                )
                .exclude(id=team.id)
                .count()
            )

            group_count = (
                Team.objects.filter(
                    college=college,
                    students=student,
                    item__item_type="group",
                )
                .exclude(id=team.id)
                .count()
            )

            if item.item_type == "single" and single_count >= single_limit:
                messages.error(
                    request,
                    f"{student.name} already reached {single_limit} single items."
                )
                return redirect("team_creation")

            if item.item_type == "group" and group_count >= group_limit:
                messages.error(
                    request,
                    f"{student.name} already reached {group_limit} group items."
                )
                return redirect("team_creation")

        team.students.set(selected_students)
        messages.success(request, f"Team '{item.name}' updated successfully.")
        return redirect("team_creation")

    return redirect("team_creation")



#🗑️ DELETE TEAM
@login_required
def delete_team(request, team_id):
    college = get_object_or_404(College, user=request.user)
    team = get_object_or_404(Team, id=team_id, college=college)
    team.delete()
    messages.success(request, "Team deleted")
    return redirect("team_creation")



#📊 STUDENT SUMMARY (FIXED)
@login_required
def student_summary(request):
    if request.user.role != "college":
        return redirect("home")

    college = get_object_or_404(College, user=request.user)

    # 🔎 GET FILTERS
    q = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    item_type = request.GET.get("item_type", "").strip()

    students = Student.objects.filter(college=college)

    # 🔍 SEARCH FILTER (student name)
    if q:
        students = students.filter(name__icontains=q)

    student_data = []

    for student in students:
        teams = student.team_set.select_related("item")

        # 🎯 CATEGORY FILTER
        if category:
            teams = teams.filter(item__category=category)

        # 🎯 ITEM TYPE FILTER
        if item_type:
            teams = teams.filter(item__item_type=item_type)

        single = teams.filter(item__item_type="single").count()
        group = teams.filter(item__item_type="group").count()

        # ❗ Skip students with no teams after filtering
        if not teams.exists():
            continue

        student_data.append({
            "student": student,
            "single": single,
            "group": group,
            "total": single + group,
            "teams": teams,
        })

    return render(request, "college/student_summary.html", {
        "student_data": student_data,
        "categories": Item.objects.values_list("category", flat=True).distinct(),
        "selected_category": category,
        "selected_type": item_type,
        "q": q,
        "college": college,
    })








#Result committee Dashboard

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Count, Sum, Q
from django.utils.timezone import now

from .models import Item, Result, Team, College
from .utils import calculate_points
from django.views.decorators.cache import never_cache


from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache

@never_cache
@login_required
def organizer_dashboard(request):
    return render(request, "organizer/dashboard.html", {
        "total_items": Item.objects.count(),
        "total_teams": Team.objects.count(),
        "total_results": Result.objects.filter(is_deleted=False).count(),
        "total_colleges": College.objects.count(),
    })


from django.db.models import Count, Q

@never_cache
@login_required
def organizer_items(request):
    q = request.GET.get("q", "")
    category = request.GET.get("category", "")
    status = request.GET.get("status", "")  # published / pending

    items = Item.objects.annotate(
        result_count=Count("results", filter=Q(results__is_deleted=False)),
        team_count=Count("team", distinct=True)
    )

    # 🔍 Search by event name
    if q:
        items = items.filter(name__icontains=q)

    # 🗂️ Filter by category
    if category:
        items = items.filter(category=category)

    # ✅ Filter by result status
    if status == "published":
        items = items.filter(result_count__gt=0)
    elif status == "pending":
        items = items.filter(result_count=0)

    total_items = items.count()
    published_items = items.filter(result_count__gt=0).count()
    pending_items = total_items - published_items

    # For dropdown
    categories = Item.objects.values_list("category", flat=True).distinct()

    return render(request, "organizer/items.html", {
        "items": items,
        "total_items": total_items,
        "published_items": published_items,
        "pending_items": pending_items,
        "categories": categories,
        "selected_category": category,
        "selected_status": status,
        "query": q,
    })


@never_cache
@login_required
def add_results(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    teams = Team.objects.filter(item=item).select_related("college")

    if request.method == "POST":
        Result.objects.filter(item=item).update(is_deleted=True)

        for team in teams:
            position = request.POST.get(f"position_{team.id}")
            grade = request.POST.get(f"grade_{team.id}")

            if position and grade:
                Result.objects.update_or_create(
                    item=item,
                    team=team,
                    defaults={
                        "position": int(position),
                        "grade": grade,
                        "points": calculate_points(int(position), grade),
                        "is_deleted": False,
                        "created_at": now(),
                    },
                )

        messages.success(request, "Results saved successfully")
        return redirect("organizer_items")

    return render(request, "organizer/add_results.html", {
        "item": item,
        "teams": teams,
    })

@never_cache
@login_required
def view_results(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    results = Result.objects.filter(item=item, is_deleted=False).select_related(
        "team__college"
    )
    return render(request, "organizer/view_results.html", {
        "item": item,
        "results": results,
    })

@never_cache
@login_required
def edit_result(request, result_id):
    result = get_object_or_404(Result, id=result_id)

    if request.method == "POST":
        position = int(request.POST["position"])
        grade = request.POST["grade"]

        result.position = position
        result.grade = grade
        result.points = calculate_points(position, grade)
        result.save()

        messages.success(request, "Result updated successfully")
        return redirect("view_results", item_id=result.item.id)

    return render(request, "organizer/edit_result.html", {
        "result": result,
        "position_choices": Result.POSITION_CHOICES,   # ✅ FIX
        "grade_choices": [g[0] for g in Result.GRADE_CHOICES],  # ✅ FIX
    })


@never_cache
@login_required
def delete_item_results(request, item_id):
    Result.objects.filter(item_id=item_id).update(is_deleted=True)
    messages.warning(request, "Results deleted (Undo available)")
    return redirect("organizer_items")

@never_cache
@login_required
def undo_delete_results(request, item_id):
    Result.objects.filter(item_id=item_id).update(is_deleted=False)
    messages.success(request, "Results restored")
    return redirect("organizer_items")

@never_cache
@login_required
def college_ranking_live(request):
    rankings = (
        Result.objects.filter(is_deleted=False)
        .values("team__college__college_name")
        .annotate(total_points=Sum("points"))
        .order_by("-total_points")
    )
    return render(request, "organizer/college_ranking.html", {"rankings": rankings})


from django.db.models import Q
@never_cache
@login_required
def organizer_student_results(request):
    query = request.GET.get("q", "")
    sort_by = request.GET.get("sort", "latest")

    results = Result.objects.filter(is_deleted=False).select_related(
        "item", "team__college"
    )

    # Search
    if query:
        results = results.filter(
            Q(team__college__college_name__icontains=query) |
            Q(item__name__icontains=query)
        )

    # Sorting
    if sort_by == "category":
        results = results.order_by("item__category", "item__name")

    elif sort_by == "college":
        results = results.order_by("team__college__college_name")

    elif sort_by == "student":
        # assuming Team has student_name or similar
        results = results.order_by("team__name")  # change if needed

    else:  # latest
        results = results.order_by("-created_at")

    return render(request, "organizer/student_results.html", {
        "results": results,
        "query": query,
        "sort_by": sort_by,
    })


#overall_result_views.py
#Leaderboard
#pointTable

from django.db.models import Sum, Q
from django.shortcuts import render
from .models import Result, Item


def public_results(request):
    category = request.GET.get("category", "").strip()
    q = request.GET.get("q", "").strip()

    # ---------------------------
    # Overall College Leaderboard
    # ---------------------------
    overall_colleges = (
        Result.objects
        .filter(is_deleted=False)
        .values("team__college__college_name")
        .annotate(total_points=Sum("points"))
        .order_by("-total_points")
    )

    # ---------------------------
    # Category-wise College Leaderboard
    # ---------------------------
    category_colleges = []
    if category:
        category_colleges = (
            Result.objects
            .filter(is_deleted=False, item__category__iexact=category)
            .values("team__college__college_name")
            .annotate(total_points=Sum("points"))
            .order_by("-total_points")
        )

    # ---------------------------
    # Top Individual Performers (Single Items)
    # ---------------------------
    individual_top = (
        Result.objects
        .filter(is_deleted=False, item__item_type="single")
        .values(
            "team__students__name",
            "team__college__college_name"
        )
        .annotate(total_points=Sum("points"))
        .order_by("-total_points")[:5]
    )

    # ---------------------------
    # FULL RESULT LIST
    # ---------------------------
    full_results = (
        Result.objects
        .filter(is_deleted=False)
        .select_related("item", "team__college")
        .prefetch_related("team__students")
        .order_by("-created_at")
    )

    # Search
    if q:
        full_results = full_results.filter(
            Q(item__name__icontains=q) |
            Q(item__category__icontains=q) |
            Q(team__college__college_name__icontains=q) |
            Q(team__students__name__icontains=q)
        )

    categories = Item.objects.values_list("category", flat=True).distinct()

    # ---------------------------
    # AJAX LIVE REFRESH
    # ---------------------------
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return render(request, "results_live.html", {
            "overall_colleges": overall_colleges,
            "individual_top": individual_top,
            "full_results": full_results,
        })

    return render(request, "public_results.html", {
        "overall_colleges": overall_colleges,
        "category_colleges": category_colleges,
        "categories": categories,
        "selected_category": category,
        "individual_top": individual_top,
        "full_results": full_results,
        "q": q,
    })



#admin Dashboardddd
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password

from .models import College, SiteSetting

User = get_user_model()


def admin_only(user):
    return user.is_authenticated and user.role == "admin"


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not admin_only(request.user):
            messages.error(request, "Admin access only")
            return redirect("login")
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
def admin_dashboard(request):
    if not admin_only(request.user):
        return redirect("home")

    return render(request, "admin/dashboard.html")


@login_required
def admin_users(request):
    if not admin_only(request.user):
        messages.error(request, "Access denied.")
        return redirect("home")

    users = User.objects.all().order_by("role", "username")

    # ✅ FIX: College linked via user FK
    college_map = {
        college.user.username: college.college_name
        for college in College.objects.select_related("user")
    }

    for u in users:
        if u.role == "college":
            u.college_display = college_map.get(u.username, "—")
        else:
            u.college_display = "—"

    return render(request, "admin/users.html", {
        "users": users
    })


@login_required
def admin_add_user(request):
    if not admin_only(request.user):
        messages.error(request, "Access denied.")
        return redirect("home")

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        role = request.POST.get("role")
        college_name = request.POST.get("college_name", "").strip()
        district = request.POST.get("district", "").strip()

        if not username or not password or not role:
            messages.error(request, "All fields are required.")
            return redirect("admin_add_user")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect("admin_add_user")

        # ✅ CREATE USER
        user = User.objects.create_user(
            username=username,
            password=password,
            role=role
        )

        # ✅ CREATE COLLEGE PROFILE IF ROLE = COLLEGE
        if role == "college":
            if not college_name:
                user.delete()
                messages.error(request, "College name is required.")
                return redirect("admin_add_user")

            College.objects.create(
                user=user,
                college_name=college_name,
                district=district,
            )

        messages.success(request, "User created successfully.")
        return redirect("admin_users")

    return render(request, "admin/add_user.html")



#🔁 ADMIN – ACTIVATE / DEACTIVATE USER
@login_required
def admin_toggle_user(request, user_id):
    if not admin_only(request.user):
        messages.error(request, "Access denied.")
        return redirect("admin_dashboard")

    user = get_object_or_404(User, id=user_id)

    if user == request.user:
        messages.warning(request, "You cannot disable your own account.")
        return redirect("admin_users")

    user.is_active = not user.is_active
    user.save()

    status = "activated" if user.is_active else "deactivated"
    messages.success(request, f"User '{user.username}' {status} successfully.")

    return redirect("admin_users")


#⚙️ ADMIN – SITE SETTINGS
@login_required
def admin_site_settings(request):
    if not admin_only(request.user):
        return redirect("home")

    setting, _ = SiteSetting.objects.get_or_create(id=1)

    if request.method == "POST":
        setting.allow_student_registration = "allow" in request.POST
        setting.save()
        messages.success(request, "Settings updated successfully.")

    return render(request, "admin/settings.html", {"setting": setting})










@login_required
def admin_edit_user(request, user_id):
    if not admin_only(request.user):
        return redirect("home")

    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        user.username = request.POST.get("username", user.username)

        # ✅ CORRECT WAY
        if request.POST.get("password"):
            user.set_password(request.POST["password"])

        user.role = request.POST.get("role", user.role)
        user.save()

        messages.success(request, "User updated successfully.")
        return redirect("admin_users")

    return render(request, "admin/edit_user.html", {"u": user})


#Delete User

@login_required
def admin_delete_user(request, user_id):
    if not admin_only(request.user):
        messages.error(request, "Access denied.")
        return redirect("admin_dashboard")

    user = get_object_or_404(User, id=user_id)

    # ❌ Prevent self-delete
    if user == request.user:
        messages.warning(request, "You cannot delete your own account.")
        return redirect("admin_users")

    username = user.username
    user.delete()   # ✅ cascades College automatically

    messages.success(request, f"User '{username}' deleted successfully.")
    return redirect("admin_users")


# accounts/views.py

@login_required
def organizer_send_appeal_result(request):
    if request.user.role != "organizer":
        return redirect("home")


    if request.method == "POST":
        college = College.objects.get(id=request.POST["college"])
        item = Item.objects.get(id=request.POST["item"])

        status = request.POST["status"]
        position = request.POST.get("position") or None
        grade = request.POST.get("grade") or None
        message = request.POST.get("message", "")
        image = request.FILES.get("result_image")

        AppealNotification.objects.create(
            college=college,
            item=item,
            status=status,
            position=position if status == "accepted" else None,
            grade=grade if status == "accepted" else None,
            message=message,
            result_image=image,
            sent_by=request.user,
        )

        messages.success(request, "Appeal result sent to college.")
        return redirect("organizer_dashboard")

    return render(request, "organizer/send_appeal_result.html", {
        "items": Item.objects.all(),
        "colleges": College.objects.all(),
    })

@login_required
def college_inbox(request):
    if request.user.role != "college":
        return redirect("home")


    college = College.objects.get(user=request.user)

    messages_qs = AppealNotification.objects.filter(
        college=college
    ).order_by("-created_at")

    return render(request, "college/inbox.html", {
        "messages": messages_qs
    })



#   EXPORTT

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

  # ✅ REQUIRED

from accounts.models import (
    Item,
    College,
    Registration,
    Team,
)



# =====================================================
# Unified participation generator (CORE FIX)
# =====================================================
def get_teamwise_participation_rows(filters):
    """
    Returns TEAM-WISE rows:
    {
        college,
        item,
        category,
        type,
        students (list)
    }
    """
    rows = []

    # =========================
    # SINGLE EVENTS (Registration)
    # =========================
    reg_qs = Registration.objects.select_related(
        "student",
        "student__college",
        "item"
    )

    if filters.get("item"):
        reg_qs = reg_qs.filter(item_id=filters["item"])
    if filters.get("college"):
        reg_qs = reg_qs.filter(student__college_id=filters["college"])
    if filters.get("category"):
        reg_qs = reg_qs.filter(item__category=filters["category"])

    for r in reg_qs:
        rows.append({
            "college": r.student.college,
            "item": r.item,
            "category": r.item.category,
            "type": "Single",
            "students": [r.student],
        })

    # =========================
    # GROUP EVENTS (Team.students)
    # =========================
    team_qs = Team.objects.select_related(
        "college",
        "item"
    ).prefetch_related("students")

    if filters.get("item"):
        team_qs = team_qs.filter(item_id=filters["item"])
    if filters.get("college"):
        team_qs = team_qs.filter(college_id=filters["college"])
    if filters.get("category"):
        team_qs = team_qs.filter(item__category=filters["category"])

    for team in team_qs:
        rows.append({
            "college": team.college,
            "item": team.item,
            "category": team.item.category,
            "type": "Group",
            "students": list(team.students.all()),
        })

    return rows



# =====================================================
# DASHBOARD
# =====================================================
@login_required
def participation_export_dashboard(request):
    items = Item.objects.all().order_by("name")
    colleges = College.objects.all().order_by("college_name")
    categories = Item.objects.values_list(
        "category",
        flat=True
    ).distinct().order_by("category")

    data = get_teamwise_participation_rows(request.GET)

    context = {
        "items": items,
        "colleges": colleges,
        "categories": categories,
        "data": data[:200],  # preview limit
        "selected": request.GET,
        "total_count": len(data),
    }

    return render(
        request,
        "organizer/participation_exports.html",
        context
    )

# =====================================================
# EXCEL EXPORT
# =====================================================
@login_required
def export_excel(request):
    data = get_teamwise_participation_rows(request.GET)

    wb = Workbook()
    ws = wb.active
    ws.title = "Participation (Team-wise)"

    ws.append([
        "College",
        "Item",
        "Category",
        "Type",
        "Participants",
    ])

    for row in data:
        students = ", ".join(
            [s.name for s in row["students"]]
        )

        ws.append([
            row["college"].college_name,
            row["item"].name,
            row["category"],
            row["type"],
            students,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=participation_teamwise.xlsx"
    )

    wb.save(response)
    return response




@login_required
def export_pdf(request):
    data = get_teamwise_participation_rows(request.GET)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        "attachment; filename=participation_teamwise.pdf"
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        spaceAfter=4,
    )

    header_style = ParagraphStyle(
        "header",
        parent=styles["Normal"],
        fontSize=9,
        leading=11,
        fontName="Helvetica-Bold",
    )

    def P(text):
        return Paragraph(str(text), cell_style)

    # -------------------------
    # TABLE HEADER
    # -------------------------
    table_data = [[
        Paragraph("College", header_style),
        Paragraph("Item", header_style),
        Paragraph("Category", header_style),
        Paragraph("Type", header_style),
        Paragraph("Participants", header_style),
    ]]

    # -------------------------
    # TABLE ROWS
    # -------------------------
    for row in data:
        participants = "<br/>".join(
            [s.name for s in row["students"]]
        ) or "-"

        table_data.append([
            P(row["college"].college_name),
            P(row["item"].name),
            P(row["category"]),
            P(row["type"]),
            KeepInFrame(
                180,  # width
                200,  # height
                [P(participants)],
                mode="shrink",
            ),
        ])

    table = Table(
        table_data,
        colWidths=[110, 100, 80, 60, 180],
        repeatRows=1,
        hAlign="LEFT",
    )

    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))

    doc.build([table])
    return response


# =====================================================
# PDF RESULT EXPORT
# =====================================================
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from accounts.models import Result, Item, College


from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from accounts.models import Result, Item, College


@login_required
def result_export_dashboard(request):
    # =======================
    # Dropdown filter data
    # =======================
    items = Item.objects.all().order_by("name")
    colleges = College.objects.all().order_by("college_name")
    categories = (
        Item.objects.values_list("category", flat=True)
        .distinct()
        .order_by("category")
    )

    # =======================
    # Base queryset (TEAM-wise)
    # =======================
    qs = (
        Result.objects
        .select_related(
            "item",
            "team",
            "team__college",
        )
        .prefetch_related(
            "team__students"
        )
        .filter(is_deleted=False)
    )

    # =======================
    # Apply filters
    # =======================
    item_id = request.GET.get("item")
    college_id = request.GET.get("college")
    category = request.GET.get("category")

    if item_id:
        qs = qs.filter(item_id=item_id)

    if college_id:
        qs = qs.filter(team__college_id=college_id)

    if category:
        qs = qs.filter(item__category=category)

    # =======================
    # Build TEAM-WISE rows
    # =======================
    rows = []
    for result in qs:
        rows.append({
            "college": result.team.college,
            "item": result.item,
            "category": result.item.category,
            "position": result.get_position_display(),
            "grade": result.grade,
            "points": result.points,
            "students": result.team.students.all(),  # 👈 key change
        })

    # =======================
    # Context
    # =======================
    context = {
        "items": items,
        "colleges": colleges,
        "categories": categories,
        "data": rows[:200],      # preview limit
        "selected": request.GET,
        "total_count": len(rows),
    }

    return render(
        request,
        "organizer/result_exports.html",
        context,
    )





@login_required
def export_result_excel(request):
    # =======================
    # Base queryset (TEAM-wise)
    # =======================
    qs = (
        Result.objects
        .select_related("item", "team", "team__college")
        .prefetch_related("team__students")
        .filter(is_deleted=False)
    )

    # Filters
    if request.GET.get("item"):
        qs = qs.filter(item_id=request.GET["item"])
    if request.GET.get("college"):
        qs = qs.filter(team__college_id=request.GET["college"])
    if request.GET.get("category"):
        qs = qs.filter(item__category=request.GET["category"])

    # =======================
    # Workbook & Sheet
    # =======================
    wb = Workbook()
    ws = wb.active
    ws.title = "Kalolsavam Results"

    # =======================
    # Header Row
    # =======================
    headers = [
        "College",
        "Item",
        "Students",
        "Position",
        "Grade",
        "Points",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = center_align

    # =======================
    # Data Rows (TEAM-wise)
    # =======================
    for r in qs:
        students_text = ", ".join(
            s.name for s in r.team.students.all()
        ) or "—"

        ws.append([
            r.team.college.college_name,
            r.item.name,
            students_text,                  # 👈 wrapped cell
            r.get_position_display(),
            r.grade,
            r.points,
        ])

        current_row = ws.max_row
        ws.cell(row=current_row, column=1).alignment = wrap_align
        ws.cell(row=current_row, column=2).alignment = wrap_align
        ws.cell(row=current_row, column=3).alignment = wrap_align
        ws.cell(row=current_row, column=4).alignment = center_align
        ws.cell(row=current_row, column=5).alignment = center_align
        ws.cell(row=current_row, column=6).alignment = center_align

    # =======================
    # Column Widths (IMPORTANT)
    # =======================
    column_widths = {
        1: 35,  # College
        2: 28,  # Item
        3: 50,  # Students
        4: 12,  # Position
        5: 10,  # Grade
        6: 10,  # Points
    }

    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # =======================
    # Response
    # =======================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=kalolsavam_results.xlsx"
    wb.save(response)

    return response






@login_required
def export_result_pdf(request):
    # =======================
    # Query (TEAM-wise)
    # =======================
    qs = (
        Result.objects
        .select_related("item", "team", "team__college")
        .prefetch_related("team__students")
        .filter(is_deleted=False)
    )

    # Filters
    if request.GET.get("item"):
        qs = qs.filter(item_id=request.GET["item"])
    if request.GET.get("college"):
        qs = qs.filter(team__college_id=request.GET["college"])
    if request.GET.get("category"):
        qs = qs.filter(item__category=request.GET["category"])

    # =======================
    # Response + Document
    # =======================
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=kalolsavam_results.pdf"

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),   # ✅ VERY IMPORTANT
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]

    elements = []

    # =======================
    # Table Header
    # =======================
    table_data = [[
        "College",
        "Item",
        "Students",
        "Position",
        "Grade",
        "Points",
    ]]

    # =======================
    # Table Rows (WRAPPED)
    # =======================
    for r in qs:
        students_text = ", ".join(
            s.name for s in r.team.students.all()
        ) or "—"

        table_data.append([
            Paragraph(r.team.college.college_name, normal),
            Paragraph(r.item.name, normal),
            Paragraph(students_text, normal),  # ✅ wraps automatically
            r.get_position_display(),
            r.grade,
            str(r.points),
        ])

    # =======================
    # Column Widths (in points)
    # =======================
    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[180, 140, 260, 80, 60, 60],  # ✅ tuned widths
    )

    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.75, colors.black),

        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("VALIGN", (0, 0), (-1, -1), "TOP"),   # ✅ critical
        ("ALIGN", (3, 1), (-1, -1), "CENTER"),

        ("FONTSIZE", (0, 0), (-1, -1), 9),

        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)

    return response

#................PUBLIC DOC..............
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import PublicDocument


# -------------------------------
# PUBLIC PAGES
# -------------------------------
def event_kit_home(request):
    return render(request, "event_kit_home.html")


def event_kit_category(request, doc_type):
    documents = PublicDocument.objects.filter(
        document_type=doc_type,
        is_active=True
    ).order_by("-created_at")

    # Process each document
    for doc in documents:
        filename = doc.file.name.lower()
        doc.is_pdf = filename.endswith(".pdf")
        
        # Try to get file size if not available
        if not hasattr(doc, 'size') or not doc.size:
            try:
                doc.size = f"{doc.file.size / 1024:.1f} KB"  # Convert bytes to KB
            except:
                doc.size = "Unknown"
        
        # Try to get upload date if not available
        if not hasattr(doc, 'upload_date') or not doc.upload_date:
            if hasattr(doc, 'created_at'):
                doc.upload_date = doc.created_at.strftime("%Y-%m-%d")
            else:
                doc.upload_date = "Unknown"

    return render(request, "event_kit_list.html", {
        "documents": documents,
        "doc_type": doc_type
    })

# -------------------------------
# ADMIN MANAGEMENT
# -------------------------------
@login_required
def admin_documents(request):
    if request.user.role != "admin":
        return redirect("home")

    docs = PublicDocument.objects.all().order_by("-created_at")
    return render(request, "admin/event_documents.html", {"docs": docs})


@login_required
def admin_upload_document(request):
    if request.user.role != "admin":
        return redirect("home")

    if request.method == "POST":
        PublicDocument.objects.create(
            title=request.POST["title"],
            document_type=request.POST["document_type"],
            file=request.FILES["file"]
        )
        messages.success(request, "Document uploaded successfully.")
        return redirect("admin_documents")

    return render(request, "admin/upload_document.html")


@login_required
def admin_delete_document(request, pk):
    if request.user.role != "admin":
        return redirect("home")

    doc = get_object_or_404(PublicDocument, pk=pk)
    doc.delete()
    messages.success(request, "Document deleted.")
    return redirect("admin_documents")


